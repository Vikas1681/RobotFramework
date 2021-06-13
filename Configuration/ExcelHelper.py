import openpyxl
import os
strDataSetPath=str(os.getcwd())+"\DataSet\"

objWorkbook =""
objSheet =""
objTotalRows=""
objTotalColumns=""
currentRow =1
rowKey=[]
rowValue=[]
rowData={}

class ExcelHelper:
	
	def __init__(self,datasetFilePath):
		print("Inside Excel Helper Class")
		global strDataSetPath
		global objWorkbook
		global objSheet  
		strDataSetPath =strDataSetPath+datasetFilePath+".xlsx"
		objWorkbook=openpyxl.load_workbook(strDataSetPath)
		objSheet=objWorkbook.active
	
	def loadTestData(self,rowID):
		global objSheet
		global currentRow
		global rowKey
		global rowValue
		global rowData
		for row in objSheet.iter_rows():
			if row[0].value==rowID:
				for cell in sheet.iter_cols(min_row=currentRow-1,max_row=currentRow-1):
				rowKey.append(cell[0].value)
				for cell in sheet.iter_cols(min_row=currentRow,max_row=currentRow):
			currentRow=currentRow+1
		for index in len(rowKey)+1
			rowData.add(rowKey[index],rowValue[index])
		
	def dpString(self,keyName):
		global rowData
		return rowData[keyName]
		
		
		
				
				
				
	
		
		
			
		
		
		
		